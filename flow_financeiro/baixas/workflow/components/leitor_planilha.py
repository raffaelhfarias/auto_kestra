"""
Componente para leitura de planilhas .xls e .xlsx de baixas.
Le as guias relevantes e extrai produtos (coluna A) e quantidades (coluna C).
- .xls  -> lido com xlrd
- .xlsx -> lido com openpyxl
"""

import os
import logging

logger = logging.getLogger(__name__)

# Guias validas que devem ser processadas e seu mapeamento para o motivo no sistema
GUIAS_VALIDAS = {
    "Avarias": "AVARIAS",
    "Brindes ou Doações": "BRINDES",
    "Demonstradores": "DEMONSTRADORES",
    "Produtos Vencidos": "PRODUTOS VENCIDOS",
}


def _ler_com_xlrd(caminho_arquivo: str) -> dict:
    """Le planilha .xls usando xlrd."""
    import xlrd

    workbook = xlrd.open_workbook(caminho_arquivo)
    resultado = {}

    for nome_guia in workbook.sheet_names():
        if nome_guia not in GUIAS_VALIDAS:
            logger.info(f"Guia '{nome_guia}' ignorada (nao eh uma guia de baixa valida).")
            continue

        sheet = workbook.sheet_by_name(nome_guia)
        produtos = []

        for row_idx in range(1, sheet.nrows):
            cell_produto = sheet.cell_value(row_idx, 0)
            cell_quantidade = sheet.cell_value(row_idx, 2)

            codigo_produto = str(cell_produto).strip()
            quantidade = str(cell_quantidade).strip()

            if codigo_produto.endswith(".0"):
                codigo_produto = codigo_produto[:-2]
            if quantidade.endswith(".0"):
                quantidade = quantidade[:-2]

            if not codigo_produto or not quantidade:
                logger.debug(f"Linha {row_idx + 1} da guia '{nome_guia}' vazia, pulando.")
                continue

            produtos.append({
                "produto": codigo_produto,
                "quantidade": quantidade,
            })

        if produtos:
            resultado[nome_guia] = produtos
            logger.info(f"Guia '{nome_guia}': {len(produtos)} produto(s) encontrado(s).")
        else:
            logger.warning(f"Guia '{nome_guia}' nao possui produtos validos.")

    return resultado


def _ler_com_openpyxl(caminho_arquivo: str) -> dict:
    """Le planilha .xlsx usando openpyxl."""
    import openpyxl

    workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
    resultado = {}

    for nome_guia in workbook.sheetnames:
        if nome_guia not in GUIAS_VALIDAS:
            logger.info(f"Guia '{nome_guia}' ignorada (nao eh uma guia de baixa valida).")
            continue

        sheet = workbook[nome_guia]
        produtos = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 3:
                continue

            cell_produto = row[0]  # Coluna A
            cell_quantidade = row[2]  # Coluna C

            if cell_produto is None or cell_quantidade is None:
                logger.debug(f"Linha {row_idx} da guia '{nome_guia}' vazia, pulando.")
                continue

            codigo_produto = str(cell_produto).strip()
            quantidade = str(cell_quantidade).strip()

            if codigo_produto.endswith(".0"):
                codigo_produto = codigo_produto[:-2]
            if quantidade.endswith(".0"):
                quantidade = quantidade[:-2]

            if not codigo_produto or not quantidade:
                continue

            produtos.append({
                "produto": codigo_produto,
                "quantidade": quantidade,
            })

        if produtos:
            resultado[nome_guia] = produtos
            logger.info(f"Guia '{nome_guia}': {len(produtos)} produto(s) encontrado(s).")
        else:
            logger.warning(f"Guia '{nome_guia}' nao possui produtos validos.")

    workbook.close()
    return resultado


def ler_planilha_baixas(caminho_arquivo: str) -> dict:
    """
    Le uma planilha .xls ou .xlsx e retorna os produtos de cada guia valida.

    Args:
        caminho_arquivo: Caminho absoluto para o arquivo .xls ou .xlsx

    Returns:
        Dicionario com chave = nome da guia e valor = lista de dicts com 'produto' e 'quantidade'.
    """
    if not os.path.exists(caminho_arquivo):
        raise FileNotFoundError(f"Arquivo nao encontrado: {caminho_arquivo}")

    logger.info(f"Abrindo planilha: {caminho_arquivo}")

    # Escolhe o leitor baseado na extensão
    ext = os.path.splitext(caminho_arquivo)[1].lower()

    if ext == ".xls":
        resultado = _ler_com_xlrd(caminho_arquivo)
    elif ext == ".xlsx":
        resultado = _ler_com_openpyxl(caminho_arquivo)
    else:
        logger.error(f"Extensao '{ext}' nao suportada. Use .xls ou .xlsx.")
        return {}

    if not resultado:
        logger.warning("Nenhuma guia valida com produtos encontrada na planilha.")

    return resultado
